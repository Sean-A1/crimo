import re
from datetime import date
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from prediction.models import MlbPredClass, normalize_code

class Command(BaseCommand):
    help = "MLB 분류 예측 결과(xlsx)를 임포트합니다 (baseball_data/2025/web_data/_YYYYMMDD-YYYYMMDD_pred.xlsx)."

    FILE_PATTERN = re.compile(r"^_([0-9]{6})-([0-9]{6})_pred\.xlsx$")
    SHEET_NAME = "predictions"

    def add_arguments(self, parser):
        parser.add_argument("--path", help="파일 경로를 직접 지정 (미지정 시 최신 패턴 파일 자동 선택)")

    def _to_float(self, v):
        try:
            if v in (None, ""): return None
            return float(str(v).strip())
        except Exception:
            return None

    def _to_int(self, v):
        try:
            if v in (None, ""): return None
            return int(float(str(v).strip()))
        except Exception:
            return None

    def handle(self, *args, **options):
        base_dir = Path(settings.BASE_DIR) / "baseball_data" / "2025" / "web_data"
        if not base_dir.exists():
            self.stderr.write(f"❌ 경로 없음: {base_dir}")
            return

        # 파일 선택
        file_path = options.get("path")
        if file_path:
            xlsx = Path(file_path)
            if not xlsx.exists():
                self.stderr.write(f"❌ 파일 없음: {xlsx}")
                return
        else:
            candidates = [p for p in base_dir.iterdir() if p.is_file() and self.FILE_PATTERN.match(p.name)]
            if not candidates:
                self.stderr.write("❌ 패턴에 맞는 파일이 없습니다: _YYYYMMDD-YYYYMMDD_pred.xlsx")
                return
            xlsx = max(candidates, key=lambda p: p.stat().st_mtime)

        self.stdout.write(f"📥 임포트 파일: {xlsx.name}")

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        if self.SHEET_NAME not in wb.sheetnames:
            self.stderr.write(f"❌ 시트 '{self.SHEET_NAME}' 를 찾을 수 없습니다. 시트들: {wb.sheetnames}")
            return
        ws = wb[self.SHEET_NAME]

        # 헤더
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_to_idx = {name: idx for idx, name in enumerate(header)}

        # ✅ 필수는 최소화
        required_min = ["date_str", "match_id", "proba_sigmoid"]
        missing = [r for r in required_min if r not in name_to_idx]
        if missing:
            self.stderr.write(f"❌ 누락 헤더: {missing} (headers={header})")
            return

        # 선택 컬럼
        col_iso   = name_to_idx.get("proba_isotonic")
        col_label = name_to_idx.get("label")

        latest = {}
        seen_dates = set()
        rows_read = rows_kept = 0

        for r in ws.iter_rows(min_row=2):
            rows_read += 1

            date_str_cell = r[name_to_idx["date_str"]].value
            match_id_cell = r[name_to_idx["match_id"]].value
            ps_cell       = r[name_to_idx["proba_sigmoid"]].value
            pi_cell       = r[col_iso].value   if col_iso   is not None else None
            label_cell    = r[col_label].value if col_label is not None else None

            if not date_str_cell or not match_id_cell:
                continue

            date_str = str(date_str_cell).strip()
            if len(date_str) != 6 or not date_str.isdigit():
                # YYMMDD만 허용
                continue

            try:
                yy = int(date_str[:2]); mm = int(date_str[2:4]); dd = int(date_str[4:6])
                dt = date(2000 + yy, mm, dd)
            except Exception:
                continue

            match_id = str(match_id_cell).strip().upper()
            if "@" not in match_id:
                continue
            away, home = [x.strip().upper() for x in match_id.split("@", 1)]
            away_norm = normalize_code(away)
            home_norm = normalize_code(home)

            # 확률
            ps = self._to_float(ps_cell)  # 필수
            if ps is None:
                continue  # 시그모이드 없으면 의미가 없음

            pi = self._to_float(pi_cell)     # ✅ 없어도 됨
            label = self._to_int(label_cell) # ✅ 없어도 됨 (ground truth 미반영 상태)

            key = (date_str, away, home)
            latest[key] = {
                "date_str": date_str,
                "date": dt,
                "away": away,
                "home": home,
                "away_norm": away_norm,
                "home_norm": home_norm,
                "proba_sigmoid": ps,
                "proba_isotonic": pi,  # None 허용
                "label": label,        # None 허용
            }
            seen_dates.add(date_str)
            rows_kept += 1

        if seen_dates:
            MlbPredClass.objects.filter(date_str__in=seen_dates).delete()

        objs = [MlbPredClass(**v) for v in latest.values()]
        MlbPredClass.objects.bulk_create(objs, ignore_conflicts=True)

        self.stdout.write(self.style.SUCCESS(
            f"✅ 읽은 행 {rows_read} → 저장 {len(objs)}건 (날짜 {len(seen_dates)}개)"
        ))
