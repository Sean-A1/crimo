# prediction/management/commands/import_prediction_mlb_regression.py
import re
from pathlib import Path
from datetime import datetime, date

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from prediction.models import MlbPredReg, normalize_code


class Command(BaseCommand):
    help = "MLB 회귀 예측 결과(xlsx)를 임포트합니다 (baseball_data/2025/_YYYYMMDD-YYYYMMDD_regression.xlsx)."

    FILE_PATTERN = re.compile(r"^_([0-9]{6})-([0-9]{6})_regression\.xlsx$")

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            help="파일 경로를 직접 지정 (미지정 시 최신 패턴 파일 자동 선택)",
        )

    # --------- parsing helpers ----------
    def parse_date(self, s):
        # 기대 포맷: YYYY-MM-DD
        try:
            dt = datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
            yymmdd = f"{dt.year%100:02d}{dt.month:02d}{dt.day:02d}"
            return dt, yymmdd
        except Exception:
            return None, None

    def parse_matchup(self, s):
        # 기대 포맷: "HOME vs AWAY"
        if not s or "vs" not in s:
            return None, None, None, None
        left, right = [x.strip().upper() for x in s.split("vs", 1)]
        home = left
        away = right
        return home, away, normalize_code(home), normalize_code(away)

    _inning_team_re = re.compile(r"^\s*([A-Za-z]{2,4})\s*:\s*([0-9|\s]+)\s*$")

    def parse_inning_block(self, s):
        """
        'SF: 0|1|... ||  CIN: 0|1|...' -> { 'SF': [0,1,...], 'CIN':[...]}
        """
        out = {}
        if not s:
            return out
        parts = [p.strip() for p in str(s).split("||")]
        for part in parts:
            m = self._inning_team_re.match(part)
            if not m:
                continue
            code = normalize_code(m.group(1))
            nums = [int(x) for x in m.group(2).replace(" ", "").split("|") if x != ""]
            out[code] = nums
        return out

    _total_re = re.compile(
        r"^\s*([A-Za-z]{2,4})\s+([0-9]+)\s*:\s*([0-9]+)\s*([A-Za-z]{2,4})\s*$"
    )

    def parse_total_block(self, s):
        """
        'SF 2 : 3 CIN' -> { 'SF':2, 'CIN':3 }
        """
        if not s:
            return {}
        m = self._total_re.match(str(s).strip())
        if not m:
            return {}
        left_team = normalize_code(m.group(1))
        left_score = int(m.group(2))
        right_score = int(m.group(3))
        right_team = normalize_code(m.group(4))
        return {left_team: left_score, right_team: right_score}

    def join_semicolons(self, nums):
        return ";".join(str(int(x)) for x in (nums or []))

    # --------- main ----------
    def handle(self, *args, **options):
        base_dir = Path(settings.BASE_DIR) / "baseball_data" / "2025"
        if not base_dir.exists():
            self.stderr.write(f"❌ 경로 없음: {base_dir}")
            return

        # 파일 선택
        file_path = options.get("path")
        if file_path:
            xlsx = Path(file_path)
            if not xlsx.exists():
                self.stderr.write(f"❌ 파일 없음: {xlsx}")
                return
        else:
            candidates = [
                p for p in base_dir.iterdir()
                if p.is_file() and self.FILE_PATTERN.match(p.name)
            ]
            if not candidates:
                self.stderr.write("❌ 패턴에 맞는 파일이 없습니다: _YYYYMMDD-YYYYMMDD_regression.xlsx")
                return
            xlsx = max(candidates, key=lambda p: p.stat().st_mtime)

        self.stdout.write(f"📥 임포트 파일: {xlsx.name}")

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active  # 첫 시트

        # 헤더 매핑
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_to_idx = {str(name).strip(): idx for idx, name in enumerate(header)}

        required = [
            "date",
            "matchup",
            "actual_inning_scores",
            "predicted_inning_scores",
            "actual_total",
            "predicted_total",
            
            "actual_starters", "predicted_starters",
            "pitching_changes", "game_pk",
        ]
        for r in required:
            if r not in name_to_idx:
                self.stderr.write(f"❌ 헤더에 '{r}' 컬럼이 없습니다. (headers={header})")
                return

        # 새 필드(없어도 통과)
        col_actual_starters = name_to_idx.get("actual_starters")
        col_pred_starters = name_to_idx.get("predicted_starters")
        col_pitching_changes = name_to_idx.get("pitching_changes")
        col_game_pk = name_to_idx.get("game_pk")
        
        seen_dates = set()
        latest = {}
        rows_read = 0

        for row in ws.iter_rows(min_row=2):
            rows_read += 1

            date_cell = row[name_to_idx["date"]].value
            matchup_cell = row[name_to_idx["matchup"]].value
            act_inn_cell = row[name_to_idx["actual_inning_scores"]].value
            pred_inn_cell = row[name_to_idx["predicted_inning_scores"]].value
            act_tot_cell = row[name_to_idx["actual_total"]].value
            pred_tot_cell = row[name_to_idx["predicted_total"]].value
            
            # 새 컬럼 안전 접근
            act_starters = (row[col_actual_starters].value if col_actual_starters is not None else None)
            pred_starters = (row[col_pred_starters].value if col_pred_starters is not None else None)
            pitching_changes = (row[col_pitching_changes].value if col_pitching_changes is not None else None)
            game_pk_val = (row[col_game_pk].value if col_game_pk is not None else None)

            dt, yymmdd = self.parse_date(date_cell)
            if not dt:
                continue

            home, away, home_norm, away_norm = self.parse_matchup(matchup_cell)
            if not home or not away:
                continue

            pred_innings = self.parse_inning_block(pred_inn_cell)
            act_innings = self.parse_inning_block(act_inn_cell)
            pred_totals = self.parse_total_block(pred_tot_cell)
            act_totals = self.parse_total_block(act_tot_cell)

            # 홈/어웨이 배열/합계 추출 (없으면 빈 배열/None)
            pred_home_arr = pred_innings.get(home_norm, [])
            pred_away_arr = pred_innings.get(away_norm, [])
            act_home_arr = act_innings.get(home_norm, []) if act_innings else []
            act_away_arr = act_innings.get(away_norm, []) if act_innings else []

            pred_home_total = pred_totals.get(home_norm)
            pred_away_total = pred_totals.get(away_norm)
            act_home_total = act_totals.get(home_norm) if act_totals else None
            act_away_total = act_totals.get(away_norm) if act_totals else None
            
            # 새 컬럼 안전 접근
            act_starters_cell  = row[col_actual_starters].value if col_actual_starters is not None else None
            pred_starters_cell = row[col_pred_starters].value  if col_pred_starters  is not None else None
            changes_cell       = row[col_pitching_changes].value if col_pitching_changes is not None else None
            game_pk_cell       = row[col_game_pk].value if col_game_pk is not None else None

            # ✅ game_pk 안전 변환 (정수/실수/문자열 모두 수용)
            game_pk = None
            if game_pk_cell is not None:
                try:
                    game_pk = str(int(float(game_pk_cell)))   # 숫자면 정수화
                except Exception:
                    v = str(game_pk_cell).strip()
                    game_pk = v or None

            # 세미콜론 형태로 저장
            rec = {
                "date": dt,
                "date_str": yymmdd,
                "home": home,
                "away": away,
                "home_norm": home_norm,
                "away_norm": away_norm,
                "pred_inn_home": self.join_semicolons(pred_home_arr),
                "pred_inn_away": self.join_semicolons(pred_away_arr),
                "act_inn_home": self.join_semicolons(act_home_arr) if act_home_arr else None,
                "act_inn_away": self.join_semicolons(act_away_arr) if act_away_arr else None,
                "pred_total_home": int(pred_home_total) if pred_home_total is not None else None,
                "pred_total_away": int(pred_away_total) if pred_away_total is not None else None,
                "act_total_home": int(act_home_total) if act_home_total is not None else None,
                "act_total_away": int(act_away_total) if act_away_total is not None else None,
                
                # 저장
                "actual_starters":    str(act_starters_cell).strip()  if act_starters_cell  else None,
                "predicted_starters": str(pred_starters_cell).strip() if pred_starters_cell else None,
                "pitching_changes":   str(changes_cell).strip()       if changes_cell       else None,
                "game_pk":            game_pk,   # ✅ 여기서 이제 정의되어 있음
            }

            key = (yymmdd, away_norm, home_norm)
            latest[key] = rec
            seen_dates.add(yymmdd)

        # 날짜 단위로 선삭제 후 저장
        if seen_dates:
            MlbPredReg.objects.filter(date_str__in=seen_dates).delete()

        objs = [MlbPredReg(**v) for v in latest.values()]
        MlbPredReg.objects.bulk_create(objs, ignore_conflicts=True)

        self.stdout.write(self.style.SUCCESS(
            f"✅ 읽은 행 {rows_read} → 저장 {len(objs)}건 (날짜 {len(seen_dates)}개)"
        ))
