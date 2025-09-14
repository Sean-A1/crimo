# leaderboards/management/commands/import_mlb_player_stats.py
import re
from pathlib import Path
from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from leaderboards.models import PlayerStatMLB
from prediction.models import normalize_code

HEAD_MAP = {
    "Rank": "rank",
    "Player": "player",
    "Age": "age",
    "Team": "team",
    "Lg": "lg",
    "WAR": "WAR",
    "G": "G",
    "PA": "PA",
    "AB": "AB",
    "R": "R",
    "H": "H",
    "2B": "_2B",
    "3B": "_3B",
    "HR": "HR",
    "RBI": "RBI",
    "SB": "SB",
    "CS": "CS",
    "BB": "BB",
    "SO": "SO",
    "BA": "BA",
    "OBP": "OBP",
    "SLG": "SLG",
    "OPS": "OPS",
    "OPS+": "OPS_plus",
    "rOBA": "rOBA",
    "Rbat+": "Rbat_plus",
    "TB": "TB",
    "GIDP": "GIDP",
    "HBP": "HBP",
    "SH": "SH",
    "SF": "SF",
    "IBB": "IBB",
    "Pos": "Pos",
    "Awards": "Awards",
}

def to_int(v):
    try:
        if v in (None, ""): return None
        return int(float(str(v).strip()))
    except Exception:
        return None

def to_float(v):
    try:
        if v in (None, ""): return None
        return float(str(v).strip())
    except Exception:
        return None

class Command(BaseCommand):
    help = "MLB 선수 스탯(2025 / BRef WAR 리그 시트)을 임포트합니다."

    def add_arguments(self, parser):
        parser.add_argument("--season", type=int, default=2025)
        parser.add_argument("--path", help="엑셀 경로 직접 지정")

    def handle(self, *args, **opts):
        season = opts["season"]
        if opts.get("path"):
            xlsx = Path(opts["path"])
        else:
            xlsx = Path(settings.BASE_DIR) / "baseball_data" / "2025" / "MLB_2025_BRef_WAR_ByLeague.xlsx"

        if not xlsx.exists():
            self.stderr.write(f"❌ 파일을 찾을 수 없음: {xlsx}")
            return

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active

        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_to_idx = {h: i for i, h in enumerate(header)}

        # 필수 헤더 확인
        required = ["Rank", "Player", "Team", "Lg", "WAR"]
        missing = [h for h in required if h not in name_to_idx]
        if missing:
            self.stderr.write(f"❌ 누락 헤더: {missing} (headers={header})")
            return

        upserts = 0
        total = 0

        for row in ws.iter_rows(min_row=2):
            total += 1
            raw = {h: (row[i].value if h in name_to_idx else None) for h, i in name_to_idx.items()}

            team = (raw.get("Team") or "").strip()
            team_norm = normalize_code(team)

            data = {
                "season": season,
                "league": "mlb",
                "team": team,
                "team_norm": team_norm,
                "lg": (raw.get("Lg") or "").strip(),
                "player": (raw.get("Player") or "").strip(),
                "rank": to_int(raw.get("Rank")),
                "age": to_int(raw.get("Age")),
                "WAR": to_float(raw.get("WAR")),
                "G": to_int(raw.get("G")),
                "PA": to_int(raw.get("PA")),
                "AB": to_int(raw.get("AB")),
                "R": to_int(raw.get("R")),
                "H": to_int(raw.get("H")),
                "_2B": to_int(raw.get("2B")),
                "_3B": to_int(raw.get("3B")),
                "HR": to_int(raw.get("HR")),
                "RBI": to_int(raw.get("RBI")),
                "SB": to_int(raw.get("SB")),
                "CS": to_int(raw.get("CS")),
                "BB": to_int(raw.get("BB")),
                "SO": to_int(raw.get("SO")),
                "BA": to_float(raw.get("BA")),
                "OBP": to_float(raw.get("OBP")),
                "SLG": to_float(raw.get("SLG")),
                "OPS": to_float(raw.get("OPS")),
                "OPS_plus": to_int(raw.get("OPS+")),
                "rOBA": to_float(raw.get("rOBA")),
                "Rbat_plus": to_int(raw.get("Rbat+")),
                "TB": to_int(raw.get("TB")),
                "GIDP": to_int(raw.get("GIDP")),
                "HBP": to_int(raw.get("HBP")),
                "SH": to_int(raw.get("SH")),
                "SF": to_int(raw.get("SF")),
                "IBB": to_int(raw.get("IBB")),
                "Pos": (str(raw.get("Pos")).strip() if raw.get("Pos") else None),
                "Awards": (str(raw.get("Awards")).strip() if raw.get("Awards") else None),
            }

            # 키: (season, league, player, team_norm)
            obj, created = PlayerStatMLB.objects.update_or_create(
                season=season,
                league="mlb",
                player=data["player"],
                team_norm=team_norm,
                defaults=data,
            )
            upserts += 1

        self.stdout.write(self.style.SUCCESS(f"✅ 총 {total}행 → upsert {upserts}건 완료 (season={season})"))
